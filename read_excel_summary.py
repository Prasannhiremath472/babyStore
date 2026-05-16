import openpyxl

wb = openpyxl.load_workbook("BABY SHOP STOCK PACHBANGLA 2026 - Copy.xlsx")
print("=" * 80)
print("EXCEL FILE SUMMARY")
print("=" * 80)
print(f"Total Sheets: {len(wb.sheetnames)}")
print(f"Sheet Names: {wb.sheetnames}")
print("=" * 80)

for sheet_name in wb.sheetnames:
    print(f"\n\nSHEET: {sheet_name}")
    print("-" * 80)
    ws = wb[sheet_name]
    
    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"Dimensions: {max_row} rows x {max_col} columns")
    print("-" * 80)
    
    # Print all rows
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(f"Row {i}: {row}")

