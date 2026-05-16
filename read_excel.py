import openpyxl

wb = openpyxl.load_workbook("BABY SHOP STOCK PACHBANGLA 2026 - Copy.xlsx")
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    print("\n\n=== Sheet:", sheet_name, "===")
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print("Row", i, ":", row)
